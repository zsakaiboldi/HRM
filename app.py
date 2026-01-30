import pandas as pd
import os
from pathlib import Path
from tkinter import Tk, Button, Label, filedialog, messagebox, StringVar
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelConverterApp:
    def __init__(self, master):
        self.master = master
        master.title("BMW Excel converter, hogy egyen a család")
        master.geometry("400x250")

        self.status_var = StringVar()
        self.status_var.set("Válaszd ki az Excel fájlt!")

        Label(master, text="Excel Feldolgozó App", font=("Arial", 14, "bold")).pack(pady=10)
        
        self.label = Label(master, textvariable=self.status_var, wraplength=350, fg="blue")
        self.label.pack(pady=10)

        self.browse_button = Button(master, text="1. Tallózás", command=self.browse_file, width=25, bg="#f0f0f0")
        self.browse_button.pack(pady=5)

        self.convert_button = Button(master, text="2. Konvertálás", command=self.convert_file, width=25, state="disabled", bg="#e1e1e1")
        self.convert_button.pack(pady=5)

        self.selected_file = None

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.selected_file = filename
            self.status_var.set(f"Kiválasztva: {os.path.basename(filename)}")
            self.convert_button.config(state="normal", bg="#90ee90") # Bekapcsoljuk és zöld lesz

    def process_logic(self, input_path, output_path):
        # Alap logika
        df = pd.read_excel(input_path, skiprows=5, header=None)
        df_cleaned = df.dropna(subset=[1]).copy()

        def get_hours(row):
            try:
                start = pd.to_datetime(str(row[2]), format='%H:%M')
                end = pd.to_datetime(str(row[3]), format='%H:%M')
                return (end - start).total_seconds() / 3600
            except:
                return row[8]

        df_cleaned[8] = df_cleaned.apply(get_hours, axis=1)
        final_df = df_cleaned[[0, 1, 2, 3, 4, 8]].copy()

        # Matek a Total sorhoz
        total_sum = pd.to_numeric(final_df[8], errors='coerce').sum()
        total_row = pd.DataFrame({0: [None], 1: [None], 2: [None], 3: [None], 4: ["Total"], 8: [total_sum]})
        final_df = pd.concat([final_df, total_row], ignore_index=True)

        # Mentés
        final_df.to_excel(output_path, index=False, header=False)

        # Formázás (Szürke fejléc + Oszlopszélesség)
        wb = load_workbook(output_path)
        ws = wb.active
        gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for cell in ws[1]:
            cell.fill = gray_fill
        
        max_length = 0
        for cell in ws['E']:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions['E'].width = max_length + 2
        wb.save(output_path)

    def convert_file(self):
        try:
            original_name = os.path.basename(self.selected_file)
            downloads_path = Path.home() / "Downloads" / original_name

            # Régi fájl törlése
            if downloads_path.exists():
                downloads_path.unlink()

            self.process_logic(self.selected_file, str(downloads_path))

            # Siker üzenet a UI-on
            self.status_var.set(f"Siker! A fájl a Letöltések mappába került.")
            messagebox.showinfo("Kész", "A konvertálás sikeresen lezajlott!")
            
        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt: {e}")

if __name__ == "__main__":
    root = Tk()
    app = ExcelConverterApp(root)
    root.mainloop()